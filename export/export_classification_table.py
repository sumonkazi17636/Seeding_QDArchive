"""
export/export_classification_table.py
Part 2, Step 4c: extract the results table required for submission.

Columns: repository_id, project_type, project_title, primary_class,
         secondary_class, no_project_files

Usage:
    python export/export_classification_table.py
    python export/export_classification_table.py \
        --db 23293505-sq26-classification.db \
        --out export/23293505-sq26-classification.xlsx
"""

import argparse
import pathlib
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
DB_DEFAULT = str(ROOT / "23293505-sq26-classification.db")
OUT_DEFAULT = str(ROOT / "export" / "23293505-sq26-classification.xlsx")

COLUMNS = [
    "repository_id", "project_type", "project_title",
    "primary_class", "secondary_class", "no_project_files",
]


def export_table(db_path: str, out_path: str) -> None:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    rows = conn.execute("""
        SELECT
            p.repository_id                AS repository_id,
            p.type                          AS project_type,
            p.title                         AS project_title,
            pc.primary_class_name           AS primary_class,
            pc.secondary_class_name         AS secondary_class,
            (SELECT COUNT(*) FROM FILES f WHERE f.project_id = p.id) AS no_project_files
        FROM PROJECTS p
        LEFT JOIN PROJECT_CLASSIFICATION pc ON pc.project_id = p.id
        ORDER BY p.repository_id, p.id
    """).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "classification"

    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B9C4D2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    for r in rows:
        ws.append([r[c] for c in COLUMNS])

    # borders + wrapping for body
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(COLUMNS)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [14, 15, 60, 42, 42, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    out = pathlib.Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[OK] Wrote {len(rows)} rows -> {out}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", default=DB_DEFAULT)
    parser.add_argument("--out", default=OUT_DEFAULT)
    args = parser.parse_args()
    export_table(args.db, args.out)
